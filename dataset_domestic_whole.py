import openpyxl as xl

dataset = xl.load_workbook("SCR Modelling.xlsx")
ws=dataset.active

tofile = xl.Workbook()
ws1 = tofile.active

cols = list(ws.iter_cols(min_row=2,values_only=True))
rows = list(ws.iter_rows(min_row=2,min_col=26,values_only=True))

headrow = list(ws.iter_rows(min_row=1,min_col=26,values_only=True))[0]

ws1.append(headrow)
ctr = 0
for i in cols[0]:
    ws1.append(rows[ctr])
    ctr+=1

tofile.save("dataset_rainfall_whole.xlsx")
