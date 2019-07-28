import raindata
import pandas as pd
import openpyxl as xl

wb = xl.load_workbook("/home/karthik/Cauvery Study/Data Village-wise.xlsx")
sheets = wb.sheetnames
ws.title = "Korangala"
for i in range(15):
    data = pd.read_excel("/home/karthik/Cauvery Study/Data Village-wise.xlsx",i)
    string = sheets[i]
    if i != 0:
        ws+str(i) = wb.create_sheet(string)
    import_data = data.iloc[:,[10,11,12,17,18,19,32,33]]
    
    