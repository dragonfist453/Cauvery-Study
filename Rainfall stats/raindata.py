import random
import numpy as np
import openpyxl as xl

def raindata_randomise(rainfall=[0]*12,data=None,random_seed = None):
    '''
    This function appends the list rainfall in random order over a column in data
    Both variables are default to avoid an error when run.
    rainfall is best when it's a list of size 12
    data is Nonetype by default
    '''
    random.seed(random_seed)
    append_data = []
    for i in range(len(data[:])):
        append_data.append(rainfall[random.randint(0,11)])
    append_data = np.array(append_data)
    append_data = append_data.reshape(-1,1)
    data = np.append(data,append_data,1)
    print("Raindata was put on random\n")
    return data

def raindata_input():
    '''
    This function is for taking input of the data in the current program its running in
    Takes no parameters. Returns a list.
    '''
    rainfall = []
    for i in range(12):
        rainfall.append(input(f"Enter month {i+1} mean rainfall in mm\n"))
    return rainfall

def raindata_extract(name,year):
    '''
    This function is for extracting rainfall data by specifying year out of an excel sheet given the name as parameter
    If no name given, data will be returned as a list of 0s
    '''
    try:
        rainfall = []
        wb = xl.load_workbook(name)
        ws = wb.active
        headrow = list(ws.iter_rows(min_row=1,min_col=2,values_only=True))[0]
        cols = list(ws.iter_cols(min_row=2,values_only=True))
        ctr = 1
        for string in headrow:
            if year == string:
                rainfall.extend(list(cols[ctr]))
                break
            ctr+=1
        return rainfall
    except:
        print("Wrong name entered! Please try again")
        raindata_extract(name)
