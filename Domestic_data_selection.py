import openpyxl as xl

original_data=xl.load_workbook("Meh.xlsx")
ws=original_data.active

villages = xl.Workbook()
ws1 = villages.active
ws1.title = "Korangala"
ws2 = villages.create_sheet("Ayangeri")
ws3 = villages.create_sheet("Cherangala")
ws4 = villages.create_sheet("Donikadu")
ws5 = villages.create_sheet("Mandal")
ws6 = villages.create_sheet("Madenadu")
ws7 = villages.create_sheet("Bettatur")
ws8 = villages.create_sheet("Boothana Kadu")
ws9 = villages.create_sheet("Tantipala")
ws10 = villages.create_sheet("Mankandur, Ratemane colony")
ws11 = villages.create_sheet("Tannimane")
ws12 = villages.create_sheet("Chettimane")
ws13 = villages.create_sheet("Kandanakolli")
ws14 = villages.create_sheet("Yemmethalu")
ws15 = villages.create_sheet("Others")

cols = list(ws.iter_cols(min_row=2,values_only=True))
rows = list(ws.iter_rows(min_row=2,min_col=26,values_only=True))

headrow = list(ws.iter_rows(min_row=1,min_col=26,values_only=True))[0]

ws1.append(headrow)
ws2.append(headrow)
ws3.append(headrow)
ws4.append(headrow)
ws5.append(headrow)
ws6.append(headrow)
ws7.append(headrow)
ws8.append(headrow)
ws9.append(headrow)
ws10.append(headrow)
ws11.append(headrow)
ws12.append(headrow)
ws13.append(headrow)
ws14.append(headrow)
ws15.append(headrow)

ctr=0

for value in cols[0]:
    if value == 1:
        ws1.append(rows[ctr])
    elif value == 2:
        ws2.append(rows[ctr])
    elif value == 3:
        ws3.append(rows[ctr])
    elif value == 4:
        ws4.append(rows[ctr])
    elif value == 5:
        ws5.append(rows[ctr])
    elif value == 6:
        ws6.append(rows[ctr])
    elif value == 7:
        ws7.append(rows[ctr])
    elif value == 8:
        ws8.append(rows[ctr])
    elif value == 9:
        ws9.append(rows[ctr])
    elif value == 10:
        ws10.append(rows[ctr])
    elif value == 11:
        ws11.append(rows[ctr])
    elif value == 12:
        ws12.append(rows[ctr])
    elif value == 13:
        ws13.append(rows[ctr])
    elif value == 14:
        ws14.append(rows[ctr])
    elif value == 15:
        ws15.append(rows[ctr])
    else:
        pass
    ctr+=1

villages.save("dataset-domestic.xlsx")
